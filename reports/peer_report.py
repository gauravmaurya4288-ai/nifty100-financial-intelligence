import sys
from pathlib import Path
import sqlite3

import pandas as pd
import numpy as np

# ==========================================================
# Project Root
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

sys.path.insert(0, str(PROJECT_ROOT))

# ==========================================================
# Database
# ==========================================================

DB = PROJECT_ROOT / "db" / "nifty100.db"

conn = sqlite3.connect(DB)

print("=" * 80)
print("DAY 20 - PEER COMPARISON REPORT")
print("=" * 80)

# ==========================================================
# Load Tables
# ==========================================================

print("\nLoading Tables...\n")

financial = pd.read_sql(

    "SELECT * FROM financial_ratios",

    conn

)

peer_groups = pd.read_sql(

    "SELECT * FROM peer_groups",

    conn

)

peer_percentiles = pd.read_sql(

    "SELECT * FROM peer_percentiles",

    conn

)

print("Financial Ratios :", financial.shape)

print("Peer Groups      :", peer_groups.shape)

print("Peer Percentiles :", peer_percentiles.shape)

# ==========================================================
# Pivot Percentiles
# ==========================================================

print("\nCreating Percentile Pivot...\n")

pivot = (

    peer_percentiles

    .pivot_table(

        index=["company_id", "year"],

        columns="metric",

        values="percentile_rank"

    )

    .reset_index()

)

pivot.columns.name = None

print("Pivot Shape :", pivot.shape)

print()

print(pivot.head())

# ==========================================================
# Rename Percentile Columns
# ==========================================================

rename_map = {

    "return_on_equity_pct":
        "roe_percentile",

    "return_on_capital_employed_pct":
        "roce_percentile",

    "net_profit_margin_pct":
        "npm_percentile",

    "debt_to_equity":
        "de_percentile",

    "interest_coverage":
        "icr_percentile",

    "asset_turnover":
        "asset_turnover_percentile",

    "free_cash_flow_cr":
        "fcf_percentile",

    "revenue_cagr_5yr":
        "revenue_percentile",

    "pat_cagr_5yr":
        "pat_percentile",

    "composite_quality_score":
        "quality_percentile"

}

pivot.rename(

    columns=rename_map,

    inplace=True

)

# ==========================================================
# Merge Tables
# ==========================================================

df = financial.merge(

    pivot,

    on=["company_id", "year"],

    how="left"

)

df = df.merge(

    peer_groups,

    on="company_id",

    how="left"

)

print()

print("Merged Shape :", df.shape)

# ==========================================================
# Check Missing Peer Groups
# ==========================================================

missing = df["peer_group_name"].isna().sum()

print()

print("Missing Peer Groups :", missing)

# ==========================================================
# Sort
# ==========================================================

df = df.sort_values(

    [

        "peer_group_name",

        "company_id",

        "year"

    ]

)

print()

print(df.head())

# ==========================================================
# Output Folder
# ==========================================================

OUTPUT = PROJECT_ROOT / "output"

OUTPUT.mkdir(

    exist_ok=True

)

REPORT = OUTPUT / "peer_comparison.xlsx"

print()

print("Output File")

print(REPORT)

print()

# ==========================================================
# Generate Peer Comparison Workbook
# ==========================================================

from openpyxl import load_workbook

print("\n" + "=" * 80)
print("GENERATING PEER COMPARISON WORKBOOK")
print("=" * 80)

with pd.ExcelWriter(

    REPORT,

    engine="openpyxl"

) as writer:

    peer_list = sorted(

        df["peer_group_name"]

        .dropna()

        .unique()

    )

    print(f"Peer Groups : {len(peer_list)}\n")

    for peer in peer_list:

        peer_df = (

            df[

                df["peer_group_name"] == peer

            ]

            .copy()

            .sort_values(

                [

                    "company_id",

                    "year"

                ]

            )

        )

        columns = [

            "company_id",

            "year",

            "return_on_equity_pct",

            "return_on_capital_employed_pct",

            "net_profit_margin_pct",

            "debt_to_equity",

            "interest_coverage",

            "free_cash_flow_cr",

            "revenue_cagr_5yr",

            "pat_cagr_5yr",

            "market_cap_crore",

            "pe_ratio",

            "pb_ratio",

            "dividend_yield_pct",

            "composite_quality_score",

            "roe_percentile",

            "roce_percentile",

            "npm_percentile",

            "de_percentile",

            "icr_percentile",

            "fcf_percentile",

            "revenue_percentile",

            "pat_percentile",

            "quality_percentile",

            "is_benchmark"

        ]

        columns = [

            c

            for c in columns

            if c in peer_df.columns

        ]

        peer_df = peer_df[columns]

        peer_df.to_excel(

            writer,

            sheet_name=peer[:31],

            index=False

        )

        print(f"✓ {peer:<25} {len(peer_df)} rows")

print()

# ==========================================================
# Formatting Workbook
# ==========================================================

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

print("\n" + "=" * 80)
print("FORMATTING EXCEL WORKBOOK")
print("=" * 80)

wb = load_workbook(REPORT)

# ----------------------------------------------------------
# Styles
# ----------------------------------------------------------

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True,
    size=11
)

green_fill = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

yellow_fill = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

red_fill = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC"
)

gold_fill = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

# ----------------------------------------------------------
# Format Every Sheet
# ----------------------------------------------------------

for sheet in wb.sheetnames:

    ws = wb[sheet]

    # ------------------------------------------------------
    # Header Formatting
    # ------------------------------------------------------

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------
    # Freeze Header
    # ------------------------------------------------------

    ws.freeze_panes = "A2"

    # ------------------------------------------------------
    # Auto Width
    # ------------------------------------------------------

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3

    # ------------------------------------------------------
    # Header Dictionary
    # ------------------------------------------------------

    headers = {

        cell.value: cell.column_letter

        for cell in ws[1]

    }

    # ------------------------------------------------------
    # Percentile Columns
    # ------------------------------------------------------

    percentile_columns = [

        "roe_percentile",

        "roce_percentile",

        "npm_percentile",

        "de_percentile",

        "icr_percentile",

        "fcf_percentile",

        "revenue_percentile",

        "pat_percentile",

        "quality_percentile"

    ]

    for col_name in percentile_columns:

        if col_name not in headers:
            continue

        col = headers[col_name]

        rng = f"{col}2:{col}{ws.max_row}"

        # Green >=75

        ws.conditional_formatting.add(

            rng,

            CellIsRule(

                operator="greaterThanOrEqual",

                formula=["75"],

                fill=green_fill

            )

        )

        # Yellow 25-75

        ws.conditional_formatting.add(

            rng,

            CellIsRule(

                operator="between",

                formula=["25", "75"],

                fill=yellow_fill

            )

        )

        # Red <=25

        ws.conditional_formatting.add(

            rng,

            CellIsRule(

                operator="lessThanOrEqual",

                formula=["25"],

                fill=red_fill

            )

        )

    # ------------------------------------------------------
    # Highlight Benchmark Company
    # ------------------------------------------------------

    if "is_benchmark" in headers:

        bench_col = headers["is_benchmark"]

        for row in range(2, ws.max_row + 1):

            if ws[f"{bench_col}{row}"].value == 1:

                for cell in ws[row]:

                    cell.fill = gold_fill

print()

# ==========================================================
# Add Median Row & Summary Sheet
# ==========================================================

print("\n" + "=" * 80)
print("ADDING SUMMARY")
print("=" * 80)

from openpyxl.styles import Font

summary = wb.create_sheet("Summary", 0)

summary["A1"] = "Peer Comparison Report"
summary["A1"].font = Font(size=18, bold=True)

summary["A3"] = "Peer Group"
summary["B3"] = "Companies"

summary["A3"].fill = header_fill
summary["B3"].fill = header_fill

summary["A3"].font = header_font
summary["B3"].font = header_font

peer_counts = (

    df.dropna(subset=["peer_group_name"])

      .groupby("peer_group_name")["company_id"]

      .nunique()

      .reset_index()

)

row = 4

for _, r in peer_counts.iterrows():

    summary.cell(row=row, column=1).value = r["peer_group_name"]

    summary.cell(row=row, column=2).value = int(r["company_id"])

    row += 1

summary.cell(row=row+1, column=1).value = "Total Companies"

summary.cell(row=row+1, column=2).value = int(

    df["company_id"].nunique()

)

summary.cell(row=row+2, column=1).value = "Peer Groups"

summary.cell(row=row+2, column=2).value = int(

    peer_counts.shape[0]

)

summary.cell(row=row+3, column=1).value = "Financial Records"

summary.cell(row=row+3, column=2).value = int(

    len(df)

)

# ==========================================================
# Median Row
# ==========================================================

numeric_columns = [

    "return_on_equity_pct",

    "return_on_capital_employed_pct",

    "net_profit_margin_pct",

    "debt_to_equity",

    "interest_coverage",

    "free_cash_flow_cr",

    "market_cap_crore",

    "pe_ratio",

    "pb_ratio",

    "dividend_yield_pct",

    "composite_quality_score"

]

for sheet in wb.sheetnames:

    if sheet == "Summary":
        continue

    ws = wb[sheet]

    headers = {

        cell.value: cell.column_letter

        for cell in ws[1]

    }

    last = ws.max_row + 2

    ws[f"A{last}"] = "Median"

    ws[f"A{last}"].font = Font(

        bold=True

    )

    for metric in numeric_columns:

        if metric not in headers:
            continue

        col = headers[metric]

        ws[f"{col}{last}"] = (

            f"=MEDIAN({col}2:{col}{ws.max_row})"

        )

# ==========================================================
# Save Workbook
# ==========================================================

wb.save(REPORT)

print()

print("=" * 80)
print("REPORT GENERATED SUCCESSFULLY")
print("=" * 80)

print()

print("Workbook")

print(REPORT)

print()

print("Summary")

print(f"Financial Records : {len(df)}")

print(f"Companies         : {df['company_id'].nunique()}")

print(f"Peer Groups       : {peer_counts.shape[0]}")

print()

print("Worksheets")

for s in wb.sheetnames:

    print("✓", s)

print()



conn.close()

