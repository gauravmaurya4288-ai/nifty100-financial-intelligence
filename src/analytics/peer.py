import sys
from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np

# ==========================================================
# Project Root
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

sys.path.insert(0, str(PROJECT_ROOT))

# ==========================================================
# Database
# ==========================================================

DB = PROJECT_ROOT / "db" / "nifty100.db"

conn = sqlite3.connect(DB)

print("=" * 80)
print("DAY 18 - PEER PERCENTILE ENGINE")
print("=" * 80)

# ==========================================================
# Load Tables
# ==========================================================

print("\nLoading Tables...\n")

financial = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

peer_groups = pd.read_sql(
    "SELECT * FROM peer_groups",
    conn
)

print("Financial Ratios :", financial.shape)
print("Peer Groups      :", peer_groups.shape)

# ==========================================================
# Verify Columns
# ==========================================================

print("\nFinancial Columns\n")

print(financial.columns.tolist())

print("\nPeer Group Columns\n")

print(peer_groups.columns.tolist())

# ==========================================================
# Rename Columns (if required)
# ==========================================================

if "peer_group_name" not in peer_groups.columns:

    if "peer_group" in peer_groups.columns:

        peer_groups = peer_groups.rename(
            columns={
                "peer_group": "peer_group_name"
            }
        )

# ==========================================================
# Remove Duplicate Mapping
# ==========================================================

peer_groups = (

    peer_groups

    .drop_duplicates(

        subset="company_id"

    )

)

# ==========================================================
# Merge
# ==========================================================

df = (

    financial

    .merge(

        peer_groups,

        on="company_id",

        how="left"

    )

)

print("\nMerged Shape :", df.shape)

# ==========================================================
# Companies without Peer Group
# ==========================================================

missing = df["peer_group_name"].isna().sum()

print()

print("Companies without Peer Group :", missing)

# ==========================================================
# Keep Only Required Columns
# ==========================================================

required = [

    "company_id",

    "year",

    "peer_group_name",

    "return_on_equity_pct",

    "return_on_capital_employed_pct",

    "net_profit_margin_pct",

    "debt_to_equity",

    "interest_coverage",

    "asset_turnover",

    "free_cash_flow_cr",

    "revenue_cagr_5yr",

    "pat_cagr_5yr",

    "composite_quality_score"

]

required = [

    c

    for c in required

    if c in df.columns

]

df = df[required]

print()

print("Working Shape :", df.shape)

print()

print(df.head())

# ==========================================================
# Check Peer Groups
# ==========================================================

print()

print("=" * 80)

print("Peer Groups")

print("=" * 80)

print(

    df["peer_group_name"]

    .value_counts(

        dropna=False

    )

)

print()

print("Unique Peer Groups :", df["peer_group_name"].nunique())

print()

# ==========================================================
# Percentile Ranking Engine
# ==========================================================

print("\n" + "=" * 80)
print("CALCULATING PEER PERCENTILES")
print("=" * 80)

metrics = {

    "return_on_equity_pct": False,

    "return_on_capital_employed_pct": False,

    "net_profit_margin_pct": False,

    "debt_to_equity": True,

    "interest_coverage": False,

    "asset_turnover": False,

    "free_cash_flow_cr": False,

    "revenue_cagr_5yr": False,

    "pat_cagr_5yr": False,

    "composite_quality_score": False

}

rows = []

peer_data = df.dropna(subset=["peer_group_name"]).copy()

for peer_name, group in peer_data.groupby("peer_group_name"):

    print(f"Processing : {peer_name}")

    group = group.copy()

    for metric, reverse in metrics.items():

        if metric not in group.columns:
            continue

        values = group[metric]

        if reverse:

            ranks = 1 - values.rank(
                pct=True,
                method="average"
            )

        else:

            ranks = values.rank(
                pct=True,
                method="average"
            )

        group[f"{metric}_percentile"] = (

            ranks * 100

        ).round(2)

    for _, row in group.iterrows():

        for metric in metrics.keys():

            if metric not in group.columns:
                continue

            rows.append({

                "company_id": row["company_id"],

                "year": row["year"],

                "peer_group_name": row["peer_group_name"],

                "metric": metric,

                "metric_value": row[metric],

                "percentile_rank": row[f"{metric}_percentile"]

            })

peer_percentiles = pd.DataFrame(rows)

print()

print("Rows Generated :", len(peer_percentiles))

print()

print(peer_percentiles.head())

print()

print(peer_percentiles.shape)

# ==========================================================
# Save to SQLite
# ==========================================================

print("\n" + "=" * 80)
print("SAVING PEER PERCENTILES")
print("=" * 80)

cursor = conn.cursor()

cursor.execute("""
DROP TABLE IF EXISTS peer_percentiles
""")

cursor.execute("""
CREATE TABLE peer_percentiles (

    company_id TEXT,

    year TEXT,

    peer_group_name TEXT,

    metric TEXT,

    metric_value REAL,

    percentile_rank REAL,

    PRIMARY KEY (
        company_id,
        year,
        peer_group_name,
        metric
    )

)
""")

conn.commit()

print("Table Created Successfully")

# ----------------------------------------------------------
# Insert Data
# ----------------------------------------------------------

peer_percentiles.to_sql(

    "peer_percentiles",

    conn,

    if_exists="append",

    index=False

)

print("Rows Inserted :", len(peer_percentiles))

# ----------------------------------------------------------
# Verify
# ----------------------------------------------------------

verification = pd.read_sql("""

SELECT

COUNT(*) AS total_rows,

COUNT(DISTINCT company_id) AS companies,

COUNT(DISTINCT peer_group_name) AS peer_groups,

COUNT(DISTINCT metric) AS metrics

FROM peer_percentiles

""", conn)

print()

print("Verification")

print(verification)

# ----------------------------------------------------------
# Export CSV
# ----------------------------------------------------------

output_csv = PROJECT_ROOT / "output" / "peer_percentiles.csv"

peer_percentiles.to_csv(

    output_csv,

    index=False

)

print()

print("CSV Exported")

print(output_csv)
# ==========================================================
# Generate Peer Comparison Excel
# ==========================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

print("\n" + "=" * 80)
print("GENERATING PEER COMPARISON REPORT")
print("=" * 80)

excel_file = PROJECT_ROOT / "output" / "peer_comparison.xlsx"

# ----------------------------------------------------------
# Merge benchmark information
# ----------------------------------------------------------

peer_info = peer_groups.copy()

peer_data = peer_data.merge(

    peer_info[["company_id", "is_benchmark"]],

    on="company_id",

    how="left"

)

# ----------------------------------------------------------
# Create Excel
# ----------------------------------------------------------

with pd.ExcelWriter(

    excel_file,

    engine="openpyxl"

) as writer:

    for peer in sorted(

        peer_data["peer_group_name"].dropna().unique()

    ):

        temp = peer_data[

            peer_data["peer_group_name"] == peer

        ].copy()

        temp = temp.sort_values(

            "composite_quality_score",

            ascending=False

        )

        temp.to_excel(

            writer,

            sheet_name=peer[:31],

            index=False

        )

# ----------------------------------------------------------
# Formatting
# ----------------------------------------------------------

wb = load_workbook(excel_file)

header_fill = PatternFill(

    fill_type="solid",

    fgColor="1F4E78"

)

header_font = Font(

    color="FFFFFF",

    bold=True

)

green = PatternFill(

    fill_type="solid",

    fgColor="C6EFCE"

)

yellow = PatternFill(

    fill_type="solid",

    fgColor="FFF2CC"

)

red = PatternFill(

    fill_type="solid",

    fgColor="F4CCCC"

)

gold = PatternFill(

    fill_type="solid",

    fgColor="FFD966"

)

for sheet in wb.sheetnames:

    ws = wb[sheet]

    # ----------------------------------------
    # Header
    # ----------------------------------------

    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(

            horizontal="center"

        )

    ws.freeze_panes = "A2"

    # ----------------------------------------
    # Auto Width
    # ----------------------------------------

    for column in ws.columns:

        width = max(

            len(str(cell.value))

            if cell.value is not None

            else 0

            for cell in column

        )

        ws.column_dimensions[

            get_column_letter(

                column[0].column

            )

        ].width = width + 3

    headers = {

        cell.value: cell.column_letter

        for cell in ws[1]

    }

    # ----------------------------------------
    # Percentile Formatting
    # ----------------------------------------

    for col_name in headers:

        if col_name.endswith("_percentile"):

            col = headers[col_name]

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="greaterThanOrEqual",

                    formula=["75"],

                    fill=green

                )

            )

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="between",

                    formula=["25", "75"],

                    fill=yellow

                )

            )

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="lessThanOrEqual",

                    formula=["25"],

                    fill=red

                )

            )

    # ----------------------------------------
    # Highlight Benchmark Company
    # ----------------------------------------

    if "is_benchmark" in headers:

        bench_col = headers["is_benchmark"]

        bench_index = ws[bench_col]

        for cell in bench_index[1:]:

            if cell.value == 1:

                for c in ws[cell.row]:

                    c.fill = gold

    # ----------------------------------------
    # Median Row
    # ----------------------------------------

    last = ws.max_row + 2

    ws[f"A{last}"] = "Median"

    numeric_columns = [

        c.column_letter

        for c in ws[1]

        if isinstance(

            ws[f"{c.column_letter}2"].value,

            (int, float)

        )

    ]

    for col in numeric_columns:

        ws[f"{col}{last}"] = (

            f"=MEDIAN({col}2:{col}{ws.max_row-2})"

        )

wb.save(excel_file)

print()

print("Peer Comparison Report Generated")

print(excel_file)

print()

print("=" * 80)
print("DAY 18 COMPLETED")
print("=" * 80)

conn.close()
