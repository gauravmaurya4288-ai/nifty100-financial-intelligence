from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------
# Export Screener Results
# ----------------------------------------------------------

def export_screener(results, output_file):

    with __import__("pandas").ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for sheet_name, df in results.items():

            export_cols = [

                "company_id",
                "year",

                "return_on_equity_pct",
                "return_on_capital_employed_pct",
                "return_on_assets_pct",

                "debt_to_equity",
                "interest_coverage",

                "free_cash_flow_cr",

                "revenue_cagr_5yr",
                "pat_cagr_5yr",

                "pe_ratio",
                "pb_ratio",

                "dividend_yield_pct",

                "market_cap_crore",

                "composite_quality_score"

            ]

            cols = [

                c for c in export_cols

                if c in df.columns

            ]

            df[cols].to_excel(

                writer,

                sheet_name=sheet_name,

                index=False

            )

    workbook = load_workbook(output_file)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    green = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    yellow = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C"
    )

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        # --------------------------
        # Header Formatting
        # --------------------------

        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center"
            )

        # --------------------------
        # Freeze Header
        # --------------------------

        ws.freeze_panes = "A2"

        # --------------------------
        # Auto Width
        # --------------------------

        for column in ws.columns:

            length = max(

                len(str(cell.value))

                if cell.value is not None

                else 0

                for cell in column

            )

            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = length + 3

        # --------------------------
        # Conditional Formatting
        # --------------------------

        headers = {

            cell.value: cell.column_letter

            for cell in ws[1]

        }

        # ROE

        if "return_on_equity_pct" in headers:

            col = headers["return_on_equity_pct"]

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="greaterThan",

                    formula=["20"],

                    fill=green

                )

            )

        # Debt

        if "debt_to_equity" in headers:

            col = headers["debt_to_equity"]

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="greaterThan",

                    formula=["2"],

                    fill=red

                )

            )

        # PE

        if "pe_ratio" in headers:

            col = headers["pe_ratio"]

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="lessThan",

                    formula=["20"],

                    fill=green

                )

            )

        # Composite Score

        if "composite_quality_score" in headers:

            col = headers["composite_quality_score"]

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="greaterThan",

                    formula=["80"],

                    fill=green

                )

            )

            ws.conditional_formatting.add(

                f"{col}2:{col}{ws.max_row}",

                CellIsRule(

                    operator="between",

                    formula=["60", "80"],

                    fill=yellow

                )

            )

    workbook.save(output_file)

    print("\nExcel Report Created Successfully")

    print(output_file)