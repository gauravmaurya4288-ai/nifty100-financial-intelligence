import sys
from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np

# ==========================================================
# Project Root
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

sys.path.insert(0, str(PROJECT_ROOT))

# ==========================================================
# Database
# ==========================================================

DB = PROJECT_ROOT / "db" / "nifty100.db"

conn = sqlite3.connect(DB)

print("=" * 80)
print("DAY 21 - PORTFOLIO OPTIMIZER")
print("=" * 80)

# ==========================================================
# Load Tables
# ==========================================================

print("\nLoading Tables...\n")

financial = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

companies = pd.read_sql(
    "SELECT * FROM companies",
    conn
)

sectors = pd.read_sql(
    "SELECT * FROM sectors",
    conn
)

print("Financial Ratios :", financial.shape)
print("Companies        :", companies.shape)
print("Sectors          :", sectors.shape)

# ==========================================================
# Merge Tables
# ==========================================================

df = financial.merge(
    companies,
    on="company_id",
    how="left"
)

df = df.merge(
    sectors,
    on="company_id",
    how="left"
)

print("\nMerged Shape :", df.shape)

# ==========================================================
# Prepare Latest Record
# ==========================================================

print("\nPreparing Latest Financial Records...")

# Remove TTM records
df = df[
    ~df["year"].astype(str).str.contains("TTM", case=False, na=False)
].copy()

# Extract numeric year
df["year_num"] = (
    df["year"]
      .astype(str)
      .str.extract(r'(\d{4})')[0]
)

df["year_num"] = pd.to_numeric(
    df["year_num"],
    errors="coerce"
)

# Remove invalid years
df = df.dropna(subset=["year_num"])

# Convert to integer
df["year_num"] = df["year_num"].astype(int)

# Latest record for every company
latest = (

    df

    .sort_values("year_num")

    .groupby("company_id", as_index=False)

    .tail(1)

)

print()

print("Latest Records :", latest.shape)

# ==========================================================
# Required Columns
# ==========================================================

required = [

    "company_id",

    "company_name",

    "year",

    "broad_sector",

    "sub_sector",

    "market_cap_category",

    "market_cap_crore",

    "pe_ratio",

    "pb_ratio",

    "dividend_yield_pct",

    "return_on_equity_pct",

    "return_on_capital_employed_pct",

    "net_profit_margin_pct",

    "debt_to_equity",

    "interest_coverage",

    "free_cash_flow_cr",

    "revenue_cagr_5yr",

    "pat_cagr_5yr",

    "composite_quality_score"

]

required = [

    c

    for c in required

    if c in latest.columns

]

portfolio = latest[required].copy()

print()

print("Portfolio Dataset :", portfolio.shape)

print()

print(portfolio.head())

# ==========================================================
# Missing Values
# ==========================================================

print()

print("=" * 80)

print("Missing Values")

print("=" * 80)

print(

    portfolio.isna()

             .sum()

             .sort_values(

                 ascending=False

             )

)

# ==========================================================
# Output Folder
# ==========================================================

OUTPUT = PROJECT_ROOT / "output"

OUTPUT.mkdir(

    exist_ok=True

)

print()

print("Output Folder")

print(OUTPUT)

print()
# ==========================================================
# Clean Missing Values
# ==========================================================

print("\n" + "=" * 80)
print("PREPARING INVESTMENT SCORES")
print("=" * 80)

numeric_cols = [

    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "composite_quality_score"

]

for col in numeric_cols:

    if col in portfolio.columns:

        portfolio[col] = portfolio[col].fillna(

            portfolio[col].median()

        )

# Fill text columns

portfolio["company_name"] = portfolio["company_name"].fillna(

    portfolio["company_id"]

)

portfolio["broad_sector"] = portfolio["broad_sector"].fillna(

    "Unknown"

)

portfolio["sub_sector"] = portfolio["sub_sector"].fillna(

    "Unknown"

)

portfolio["market_cap_category"] = portfolio["market_cap_category"].fillna(

    "Unknown"

)

# ==========================================================
# Normalize Helper
# ==========================================================

def normalize(series):

    s = series.copy()

    lower = s.quantile(0.05)

    upper = s.quantile(0.95)

    s = s.clip(lower, upper)

    return ((s - lower) / (upper - lower) * 100).clip(0, 100)

# ==========================================================
# Score Components
# ==========================================================

portfolio["roe_score"] = normalize(

    portfolio["return_on_equity_pct"]

)

portfolio["roce_score"] = normalize(

    portfolio["return_on_capital_employed_pct"]

)

portfolio["margin_score"] = normalize(

    portfolio["net_profit_margin_pct"]

)

portfolio["growth_score"] = normalize(

    portfolio["revenue_cagr_5yr"]

)

portfolio["profit_score"] = normalize(

    portfolio["pat_cagr_5yr"]

)

portfolio["fcf_score"] = normalize(

    portfolio["free_cash_flow_cr"]

)

portfolio["debt_score"] = 100 - normalize(

    portfolio["debt_to_equity"]

)

portfolio["valuation_score"] = 100 - (

    normalize(portfolio["pe_ratio"]) * 0.5 +

    normalize(portfolio["pb_ratio"]) * 0.5

)

portfolio["dividend_score"] = normalize(

    portfolio["dividend_yield_pct"]

)
portfolio["quality_score"] = normalize(
    portfolio["composite_quality_score"]
)

# ==========================================================
# Final Investment Score
# ==========================================================

portfolio["investment_score"] = (

      portfolio["quality_score"] * 0.30

    + portfolio["roe_score"] * 0.10

    + portfolio["roce_score"] * 0.10

    + portfolio["growth_score"] * 0.10

    + portfolio["profit_score"] * 0.10

    + portfolio["fcf_score"] * 0.10

    + portfolio["debt_score"] * 0.10

    + portfolio["valuation_score"] * 0.05

    + portfolio["dividend_score"] * 0.05

).round(2)

# ==========================================================
# Investment Grade
# ==========================================================

def investment_grade(score):

    if score >= 90:
        return "A+"

    elif score >= 80:
        return "A"

    elif score >= 70:
        return "B+"

    elif score >= 60:
        return "B"

    elif score >= 50:
        return "C"

    return "D"

portfolio["grade"] = portfolio["investment_score"].apply(

    investment_grade

)

# ==========================================================
# Risk Category
# ==========================================================

def risk(row):

    if (

        row["debt_to_equity"] < 0.5

        and

        row["interest_coverage"] > 5

    ):

        return "Low"

    elif row["debt_to_equity"] < 1.5:

        return "Moderate"

    else:

        return "High"

portfolio["risk"] = portfolio.apply(

    risk,

    axis=1

)

# ==========================================================
# Recommendation
# ==========================================================

def recommendation(score):

    if score >= 85:
        return "Strong Buy"

    elif score >= 70:
        return "Buy"

    elif score >= 55:
        return "Accumulate"

    elif score >= 40:
        return "Hold"

    return "Avoid"

portfolio["recommendation"] = portfolio["investment_score"].apply(

    recommendation

)

# ==========================================================
# Preview
# ==========================================================

portfolio = portfolio.sort_values(

    "investment_score",

    ascending=False

)

print()

print("Top 10 Companies")

print(

    portfolio[

        [

            "company_id",

            "investment_score",

            "grade",

            "risk",

            "recommendation"

        ]

    ].head(10)

)

print()

print("Grade Distribution")

print(

    portfolio["grade"].value_counts()

)

print()

print("Recommendation Distribution")

print(

    portfolio["recommendation"].value_counts()

)

# ==========================================================
# PART 3 - PORTFOLIO ALLOCATION
# ==========================================================

print("\n" + "=" * 80)
print("GENERATING PORTFOLIO")
print("=" * 80)

# ----------------------------------------------------------
# Sort by Investment Score
# ----------------------------------------------------------

portfolio = portfolio.sort_values(
    "investment_score",
    ascending=False
).reset_index(drop=True)

# ----------------------------------------------------------
# Select Top 25
# ----------------------------------------------------------

top25 = portfolio.head(25).copy()

print("\nTop Companies Selected :", len(top25))

# ----------------------------------------------------------
# Calculate Portfolio Weight
# ----------------------------------------------------------

total_score = top25["investment_score"].sum()

top25["portfolio_weight"] = (

    top25["investment_score"]

    / total_score

) * 100

top25["portfolio_weight"] = (

    top25["portfolio_weight"]

    .round(2)

)

# ----------------------------------------------------------
# Adjust Rounding Difference
# ----------------------------------------------------------

difference = round(

    100 - top25["portfolio_weight"].sum(),

    2

)

top25.loc[0, "portfolio_weight"] += difference

# ----------------------------------------------------------
# Rank
# ----------------------------------------------------------

top25.insert(

    0,

    "rank",

    range(1, len(top25) + 1)

)

# ----------------------------------------------------------
# Allocation Summary
# ----------------------------------------------------------

print()

print("=" * 80)
print("PORTFOLIO ALLOCATION")
print("=" * 80)

print(

    top25[

        [

            "rank",

            "company_id",

            "company_name",

            "investment_score",

            "portfolio_weight",

            "grade",

            "risk",

            "recommendation"

        ]

    ]

)

print()

print("Total Weight :",

      round(

          top25["portfolio_weight"].sum(),

          2

      ),

      "%"

)

# ----------------------------------------------------------
# Sector Allocation
# ----------------------------------------------------------

print()

print("=" * 80)
print("SECTOR ALLOCATION")
print("=" * 80)

sector_summary = (

    top25

    .groupby("broad_sector")

    ["portfolio_weight"]

    .sum()

    .sort_values(

        ascending=False

    )

    .round(2)

)

print(sector_summary)

# ----------------------------------------------------------
# Market Cap Allocation
# ----------------------------------------------------------

print()

print("=" * 80)
print("MARKET CAP ALLOCATION")
print("=" * 80)

market_summary = (

    top25

    .groupby("market_cap_category")

    ["portfolio_weight"]

    .sum()

    .round(2)

)

print(market_summary)

print()

# ==========================================================
# PART 4 - EXPORT REPORTS
# ==========================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("\n" + "=" * 80)
print("EXPORTING REPORTS")
print("=" * 80)

# ----------------------------------------------------------
# Output Files
# ----------------------------------------------------------

excel_file = OUTPUT / "portfolio_recommendation.xlsx"
csv_file = OUTPUT / "top25_portfolio.csv"
summary_file = OUTPUT / "portfolio_summary.csv"

# ----------------------------------------------------------
# Export CSV
# ----------------------------------------------------------

top25.to_csv(csv_file, index=False)

# ----------------------------------------------------------
# Summary Table
# ----------------------------------------------------------

summary = pd.DataFrame({

    "Metric": [

        "Total Companies",
        "Top Portfolio",
        "Average Investment Score",
        "Maximum Score",
        "Minimum Score",
        "Total Portfolio Weight"

    ],

    "Value": [

        portfolio["company_id"].nunique(),
        len(top25),
        round(portfolio["investment_score"].mean(),2),
        round(portfolio["investment_score"].max(),2),
        round(portfolio["investment_score"].min(),2),
        round(top25["portfolio_weight"].sum(),2)

    ]

})

summary.to_csv(summary_file,index=False)

# ----------------------------------------------------------
# Excel Workbook
# ----------------------------------------------------------

with pd.ExcelWriter(

    excel_file,

    engine="openpyxl"

) as writer:

    top25.to_excel(

        writer,

        sheet_name="Portfolio",

        index=False

    )

    summary.to_excel(

        writer,

        sheet_name="Summary",

        index=False

    )

# ----------------------------------------------------------
# Formatting Workbook
# ----------------------------------------------------------

wb = load_workbook(excel_file)

header_fill = PatternFill(

    fill_type="solid",

    fgColor="1F4E78"

)

header_font = Font(

    color="FFFFFF",

    bold=True,

    size=11

)

for sheet in wb.sheetnames:

    ws = wb[sheet]

    # Header

    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(horizontal="center")

    # Freeze

    ws.freeze_panes = "A2"

    # Auto Width

    for column in ws.columns:

        length = 0

        letter = get_column_letter(column[0].column)

        for cell in column:

            try:

                length = max(

                    length,

                    len(str(cell.value))

                )

            except:

                pass

        ws.column_dimensions[letter].width = length + 3

wb.save(excel_file)

# ----------------------------------------------------------
# Statistics
# ----------------------------------------------------------

print()

print("="*80)

print("PORTFOLIO SUMMARY")

print("="*80)

print()

print("Top Portfolio Companies :",len(top25))

print("Average Score           :",round(portfolio["investment_score"].mean(),2))

print("Highest Score           :",round(portfolio["investment_score"].max(),2))

print("Lowest Score            :",round(portfolio["investment_score"].min(),2))

print("Portfolio Weight        :",round(top25["portfolio_weight"].sum(),2),"%")

print()

print("Recommendation Distribution")

print(portfolio["recommendation"].value_counts())

print()

print("Grade Distribution")

print(portfolio["grade"].value_counts())

print()

print("Sector Allocation")

print(sector_summary)

print()

print("Market Cap Allocation")

print(market_summary)

print()

print("="*80)

print("FILES GENERATED")

print("="*80)

print()

print(excel_file)

print(csv_file)

print(summary_file)

print()



conn.close()