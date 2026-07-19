# =====================================================
# EXECUTIVE PORTFOLIO REPORT GENERATOR
# Sprint 5 - Day 35
# =====================================================

# -----------------------------
# Imports
# -----------------------------

import sqlite3
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt

from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)

# -----------------------------
# Project Paths
# -----------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"

OUTPUT_DIR = PROJECT_ROOT / "output"

REPORT_DIR = OUTPUT_DIR / "portfolio_reports"

CHART_DIR = REPORT_DIR / "charts"

REPORT_DIR.mkdir(parents=True, exist_ok=True)
CHART_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------
# Database Connection
# -----------------------------

def get_connection():

    return sqlite3.connect(DB_PATH)

# -----------------------------
# Load Data
# -----------------------------

def load_data():

    conn = get_connection()

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn
    )

    sectors = pd.read_sql(
        "SELECT * FROM sectors",
        conn
    )

    financial_ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    conn.close()

    analysis = pd.read_csv(
        OUTPUT_DIR / "analysis_parsed.csv"
    )

    cashflow = pd.read_excel(
        OUTPUT_DIR / "cashflow_intelligence.xlsx"
    )

    capital = pd.read_excel(
        OUTPUT_DIR / "capital_allocation_report.xlsx"
    )

    return {

        "companies": companies,
        "sectors": sectors,
        "financial_ratios": financial_ratios,
        "analysis": analysis,
        "cashflow": cashflow,
        "capital": capital

    }

# -----------------------------
# Preview Data
# -----------------------------

def preview_data():

    print("=" * 70)
    print("EXECUTIVE PORTFOLIO REPORT")
    print("=" * 70)

    data = load_data()

    for name, df in data.items():

        print(f"{name:<20}{len(df):>6} rows")

# =====================================================
# PORTFOLIO STATISTICS
# =====================================================

def portfolio_statistics():

    data = load_data()

    companies = data["companies"].copy()
    sectors = data["sectors"].copy()
    financial_ratios = data["financial_ratios"].copy()
    cashflow = data["cashflow"].copy()
    capital = data["capital"].copy()

    print("\n" + "=" * 70)
    print("PORTFOLIO STATISTICS")
    print("=" * 70)

    # -------------------------------------------------
    # Latest Financial Ratios
    # -------------------------------------------------

    financial_ratios = (
        financial_ratios
        .sort_values("year")
        .groupby("company_id", as_index=False)
        .tail(1)
    )

    # -------------------------------------------------
    # Build Portfolio Dataset
    # -------------------------------------------------

    duplicate_columns = [
        "company_name",
        "ticker",
        "website",
        "broad_sector",
        "sub_sector",
        "market_cap_category",
        "index_weight_pct",
    ]

    for df in [financial_ratios, cashflow, capital]:

        cols = [c for c in duplicate_columns if c in df.columns]

        if cols:
            df.drop(columns=cols, inplace=True)

    portfolio = (
        companies
        .merge(sectors, on="company_id", how="left")
        .merge(financial_ratios, on="company_id", how="left")
        .merge(cashflow, on="company_id", how="left")
        .merge(capital, on="company_id", how="left")
    )

    portfolio = portfolio.loc[:, ~portfolio.columns.duplicated()]

    # -------------------------------------------------
    # Numeric Conversion
    # -------------------------------------------------

    numeric_cols = [
        "market_cap_crore",
        "roe_percentage",
        "roce_percentage",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "capital_score",
        "index_weight_pct",
    ]

    for col in numeric_cols:

        if col in portfolio.columns:

            portfolio[col] = pd.to_numeric(
                portfolio[col],
                errors="coerce"
            )

    # -------------------------------------------------
    # KPI Calculation
    # -------------------------------------------------

    stats = {

        "Total Companies":
            len(portfolio),

        "Total Sectors":
            portfolio["broad_sector"].nunique(),

        "Total Market Cap":
            round(
                portfolio["market_cap_crore"].sum(),
                2
            ),

        "Average Market Cap":
            round(
                portfolio["market_cap_crore"].mean(),
                2
            ),

        "Average ROE":
            round(
                portfolio["roe_percentage"].mean(),
                2
            ),

        "Average ROCE":
            round(
                portfolio["roce_percentage"].mean(),
                2
            ),

        "Average PE":
            round(
                portfolio["pe_ratio"].mean(),
                2
            ),

        "Average PB":
            round(
                portfolio["pb_ratio"].mean(),
                2
            ),

        "Average EV/EBITDA":
            round(
                portfolio["ev_ebitda"].mean(),
                2
            ),

        "Average Capital Score":
            round(
                portfolio["capital_score"].mean(),
                2
            )

    }

    # -------------------------------------------------
    # Leaders
    # -------------------------------------------------

    leaders = {

        "Largest Company":
            portfolio.loc[
                portfolio["market_cap_crore"].idxmax(),
                "company_name"
            ],

        "Highest ROE":
            portfolio.loc[
                portfolio["roe_percentage"].idxmax(),
                "company_name"
            ],

        "Highest ROCE":
            portfolio.loc[
                portfolio["roce_percentage"].idxmax(),
                "company_name"
            ],

        "Highest Capital Score":
            portfolio.loc[
                portfolio["capital_score"].idxmax(),
                "company_name"
            ]

    }

    # -------------------------------------------------
    # Display Statistics
    # -------------------------------------------------

    print("\nPortfolio KPIs\n")

    for key, value in stats.items():
        print(f"{key:<30}: {value}")

    print("\nPortfolio Leaders\n")

    for key, value in leaders.items():
        print(f"{key:<30}: {value}")

    return portfolio, stats, leaders


# =====================================================
# PORTFOLIO CHARTS
# =====================================================

def generate_charts(portfolio):

    print("\n" + "=" * 70)
    print("GENERATING CHARTS")
    print("=" * 70)

    plt.style.use("ggplot")

    # -------------------------------------------------
    # Sector Distribution
    # -------------------------------------------------

    sector_counts = (
        portfolio["broad_sector"]
        .value_counts()
        .sort_values(ascending=False)
    )

    plt.figure(figsize=(10, 6))

    sector_counts.plot(kind="bar")

    plt.title("Companies by Sector")
    plt.xlabel("Sector")
    plt.ylabel("Number of Companies")

    plt.xticks(rotation=45, ha="right")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "sector_distribution.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # Market Cap Distribution
    # -------------------------------------------------

    plt.figure(figsize=(8, 6))

    portfolio["market_cap_crore"].dropna().hist(
        bins=20
    )

    plt.title("Market Cap Distribution")
    plt.xlabel("Market Cap (Crore)")
    plt.ylabel("Companies")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "market_cap_distribution.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # ROE Distribution
    # -------------------------------------------------

    plt.figure(figsize=(8, 6))

    portfolio["roe_percentage"].dropna().hist(
        bins=20
    )

    plt.title("ROE Distribution")
    plt.xlabel("ROE (%)")
    plt.ylabel("Companies")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "roe_distribution.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # ROCE Distribution
    # -------------------------------------------------

    plt.figure(figsize=(8, 6))

    portfolio["roce_percentage"].dropna().hist(
        bins=20
    )

    plt.title("ROCE Distribution")
    plt.xlabel("ROCE (%)")
    plt.ylabel("Companies")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "roce_distribution.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # Capital Score Distribution
    # -------------------------------------------------

    plt.figure(figsize=(8, 6))

    portfolio["capital_score"].dropna().hist(
        bins=20
    )

    plt.title("Capital Score Distribution")
    plt.xlabel("Capital Score")
    plt.ylabel("Companies")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "capital_score_distribution.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # Top 10 Market Cap Companies
    # -------------------------------------------------

    top10 = (
        portfolio
        .nlargest(10, "market_cap_crore")
        .sort_values("market_cap_crore")
    )

    plt.figure(figsize=(10, 6))

    plt.barh(
        top10["company_name"],
        top10["market_cap_crore"]
    )

    plt.title("Top 10 Companies by Market Cap")
    plt.xlabel("Market Cap (Crore)")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "top10_market_cap.png",
        dpi=300
    )

    plt.close()

    # -------------------------------------------------
    # Sector Allocation Pie Chart
    # -------------------------------------------------

    plt.figure(figsize=(8, 8))

    sector_counts.plot(
        kind="pie",
        autopct="%1.1f%%"
    )

    plt.ylabel("")

    plt.title("Portfolio Sector Allocation")

    plt.tight_layout()

    plt.savefig(
        CHART_DIR / "sector_allocation.png",
        dpi=300
    )

    plt.close()

    print("\nCharts Generated Successfully")

    print(f"Location : {CHART_DIR}")
    
# =====================================================
# EXCEL REPORT GENERATOR
# =====================================================

def generate_excel_report(portfolio, stats, leaders):

    print("\n" + "=" * 70)
    print("GENERATING EXCEL REPORT")
    print("=" * 70)

    wb = Workbook()

    # =================================================
    # Executive Summary
    # =================================================

    ws = wb.active
    ws.title = "Executive Summary"

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    ws["A1"] = "NIFTY100 EXECUTIVE PORTFOLIO REPORT"
    ws["A1"].font = title_font

    ws.append([])
    ws.append(["Portfolio KPI", "Value"])

    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill

    for key, value in stats.items():
        ws.append([key, value])

    ws.append([])
    ws.append(["Portfolio Leaders", "Company"])

    for cell in ws[15]:
        cell.font = header_font
        cell.fill = header_fill

    for key, value in leaders.items():
        ws.append([key, value])

    # =================================================
    # Portfolio Data
    # =================================================

    portfolio_sheet = wb.create_sheet("Portfolio")

    for row in portfolio.itertuples(index=False):
        portfolio_sheet.append(row)

    for col_num, column in enumerate(portfolio.columns, start=1):
        portfolio_sheet.cell(row=1, column=col_num).value = column
        portfolio_sheet.cell(row=1, column=col_num).font = header_font
        portfolio_sheet.cell(row=1, column=col_num).fill = header_fill

    # =================================================
    # Sector Summary
    # =================================================

    sector_sheet = wb.create_sheet("Sector Summary")

    sector_summary = (
        portfolio.groupby("broad_sector")
        .agg(
            Companies=("company_id", "count"),
            Avg_MarketCap=("market_cap_crore", "mean"),
            Avg_ROE=("roe_percentage", "mean"),
            Avg_ROCE=("roce_percentage", "mean"),
            Avg_PE=("pe_ratio", "mean"),
            Avg_CapitalScore=("capital_score", "mean")
        )
        .round(2)
        .reset_index()
    )

    sector_sheet.append(list(sector_summary.columns))

    for cell in sector_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in sector_summary.itertuples(index=False):
        sector_sheet.append(row)

    # =================================================
    # Rankings
    # =================================================

    ranking_sheet = wb.create_sheet("Top Companies")

    top10 = portfolio.nlargest(
        10,
        "market_cap_crore"
    )[
        [
            "company_name",
            "broad_sector",
            "market_cap_crore",
            "roe_percentage",
            "roce_percentage",
            "capital_score"
        ]
    ]

    ranking_sheet.append(list(top10.columns))

    for cell in ranking_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in top10.itertuples(index=False):
        ranking_sheet.append(row)

    # =================================================
    # Auto Width
    # =================================================

    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(length + 3, 40)

    # =================================================
    # Save Workbook
    # =================================================

    output_file = REPORT_DIR / "portfolio_summary.xlsx"

    wb.save(output_file)

    print(f"\nExcel Report Saved : {output_file}")    

# =====================================================
# EXECUTIVE PDF REPORT
# =====================================================

def generate_pdf_report(portfolio, stats, leaders):

    print("\n" + "=" * 70)
    print("GENERATING PDF REPORT")
    print("=" * 70)

    

    pdf_file = REPORT_DIR / "Executive_Portfolio_Report.pdf"

    doc = SimpleDocTemplate(str(pdf_file))

    styles = getSampleStyleSheet()

    story = []

    # =================================================
    # Title
    # =================================================

    story.append(
        Paragraph(
            "<b><font size=20>NIFTY100 Executive Portfolio Report</font></b>",
            styles["Title"]
        )
    )

    story.append(Spacer(1, 20))

    story.append(
        Paragraph(
            "Generated using the Nifty100 Financial Intelligence Platform.",
            styles["Normal"]
        )
    )

    story.append(Spacer(1, 20))

    # =================================================
    # Portfolio KPIs
    # =================================================

    story.append(
        Paragraph("<b>Portfolio Statistics</b>", styles["Heading2"])
    )

    table_data = [["Metric", "Value"]]

    for k, v in stats.items():
        table_data.append([k, str(v)])

    table = Table(table_data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 8)

    ]))

    story.append(table)

    story.append(Spacer(1, 20))

    # =================================================
    # Portfolio Leaders
    # =================================================

    story.append(
        Paragraph("<b>Portfolio Leaders</b>", styles["Heading2"])
    )

    leader_table = [["Category", "Company"]]

    for k, v in leaders.items():
        leader_table.append([k, str(v)])

    table = Table(leader_table)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke)

    ]))

    story.append(table)

    story.append(Spacer(1, 20))

    # =================================================
    # Sector Summary
    # =================================================

    story.append(
        Paragraph("<b>Sector Summary</b>", styles["Heading2"])
    )

    sector_summary = (
        portfolio.groupby("broad_sector")
        .agg(
            Companies=("company_id", "count"),
            Avg_ROE=("roe_percentage", "mean"),
            Avg_ROCE=("roce_percentage", "mean"),
            Avg_Capital=("capital_score", "mean")
        )
        .round(2)
        .reset_index()
    )

    sector_table = [list(sector_summary.columns)]

    for row in sector_summary.values.tolist():
        sector_table.append(row)

    table = Table(sector_table)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige)

    ]))

    story.append(table)

    story.append(Spacer(1, 20))

    # =================================================
    # Charts
    # =================================================

    story.append(
        Paragraph("<b>Portfolio Charts</b>", styles["Heading2"])
    )

    chart_files = [

        "sector_distribution.png",

        "market_cap_distribution.png",

        "roe_distribution.png",

        "top10_market_cap.png"

    ]

    for chart in chart_files:

        chart_path = CHART_DIR / chart

        if chart_path.exists():

            story.append(
                Image(
                    str(chart_path),
                    width=450,
                    height=280
                )
            )

            story.append(Spacer(1, 15))

    # =================================================
    # Footer
    # =================================================

    story.append(
        Paragraph(
            "<b>End of Report</b>",
            styles["Heading2"]
        )
    )

    doc.build(story)

    print(f"\nPDF Report Saved : {pdf_file}")

# -----------------------------
# Main
# -----------------------------

if __name__ == "__main__":

    preview_data()

    portfolio, stats, leaders = portfolio_statistics()

    generate_charts(portfolio)

    generate_excel_report(
    portfolio,
    stats,
    leaders
    )

    generate_pdf_report(
        portfolio,
        stats,
        leaders
    )